import os
import time
import openpyxl
import pandas as pd
from loguru import logger
from datetime import datetime
from dotenv import load_dotenv
from config import conn

class Estoques:
    def __init__(self):
        load_dotenv()
        self.path_dados = os.getenv('DATA_DIRECTORY')
        self.unid_codigos = ['001', '002', ['003','010']]
        self.conn = conn.DatabaseConnection.get_db_engine(self)
    
    def estoques_query(self, conn, unid_codigo):
        if isinstance(unid_codigo, list):
            unid_values = ",".join([f"'{code}'" for code in unid_codigo])
        else:
            unid_values = f"'{unid_codigo}'"
        query = (f"""
            SELECT 
                TO_CHAR(prun.prun_estoque1, '000000000000999D9999') AS estoque,
                TO_CHAR((prun.prun_estoque1 * prod.prod_pesoliq), '000000000000999D9999') AS qtde,
                prun.prun_unid_codigo AS unidade,
                prun.prun_ativo as tipo,
                prun.prun_prod_codigo AS prod_codigo,
                prod.prod_codbarras AS cod_barras,
                prun.prun_emb AS embalagem,
                prod.prod_marca AS marca,
                prod.prod_codigo AS cod_prod
            FROM produn AS prun 
            LEFT JOIN produtos AS prod ON prun.prun_prod_codigo = prod.prod_codigo
            WHERE prun.prun_bloqueado = 'N' 
            AND prun.prun_unid_codigo IN ({unid_values})
            AND prun.prun_ativo = 'S'
            AND prod.prod_marca IN ('BRF', 'BRF IN NATURA')
            AND prun.prun_estoque1 > 0
        """)
        return pd.read_sql_query(query, conn)

    def process_rows(self, df, unid_codigo):
        processed_rows = []
        for index, row in df.iterrows():
            caracter_adc = 'E'
            
            if unid_codigo == '001':
                cnpj_unid = '81894255000147'
            elif unid_codigo == '002':
                cnpj_unid = '81894255000228'
            else:
                cnpj_unid = '81894255000309'
                
            cod_barras = row['cod_barras'].zfill(13)
            estoque = row['qtde'].strip()
            embalagem = row['embalagem']
            if embalagem != 'KG':
                tipoUnid = "0001"
            else:
                tipoUnid = "0002"
            
            espaco_branco1 = ' '
            espaco_branco2 = ' '*20
            espaco_branco3 = ' '*8
            
            processed_row = (f'{caracter_adc}{cnpj_unid}{cod_barras}{espaco_branco1}{estoque}{espaco_branco2}{espaco_branco3}{tipoUnid}')
            processed_rows.append(processed_row)
        logger.info('Query estoques OK!')
        logger.info('Processamento de dados estoques OK!')
        return processed_rows
    
    def save_to_excel_and_txt(self, processed_rows, unid_codigo, data_atual):
        wb = openpyxl.Workbook()
        ws = wb.active
        data_atual_estoque = datetime.now().strftime("%Y%m%d")
        ws['A1'] = (f'HESTOQ1101838723010513{data_atual_estoque}')
        for index, row_value in enumerate(processed_rows, start=2):
            ws.cell(row=index, column=1).value = row_value
        
        if unid_codigo == ['003','010']:
            unid_codigo = '003'
        nome_arquivo = (f'ESTOQUESDUSNEI{unid_codigo}{data_atual}')
        ws.title = data_atual
        data_pasta = datetime.now().strftime("%Y-%m-%d")
        diretorio = f'{self.path_dados}/{data_pasta}'
        if not os.path.exists(diretorio):
                os.mkdir(diretorio)
        local_arquivo = os.path.join(f'{diretorio}/{nome_arquivo}.xlsx')
        wb.save(local_arquivo)
        
        time.sleep(2)
        
        local_arquivo_txt = os.path.join(f'{diretorio}/{nome_arquivo}.txt')
        with open(local_arquivo_txt, 'w') as txt_file:
            txt_file.write((f'HESTOQ1101838723010513{data_atual_estoque}') + '\n')
            for row in processed_rows:
                txt_file.write(row + '\n')
        
        logger.info('Arquivo estoques OK!')
    
    def estoques(self):
        for unid_codigo in self.unid_codigos:
            df = pd.concat([self.estoques_query( self.conn, unid_codigo)])
            
            processed_rows = self.process_rows(df, unid_codigo)
            data_atual = datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]
            self.save_to_excel_and_txt(processed_rows, unid_codigo, data_atual)
        
        logger.info('Funções estoques OK!')
            
if __name__ == "__main__":
    db_clientes = Estoques()    
    db_clientes.estoques()
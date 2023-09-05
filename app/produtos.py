# flake8: noqa W293
# pylint: disable=all
import os
import time
import openpyxl
import pandas as pd
from loguru import logger
from datetime import datetime
from dotenv import load_dotenv
from conn import DatabaseConnection


class Produtos:
    def __init__(self):
        load_dotenv()
        self.path_dados = os.getenv('DATA_DIRECTORY')
        self.unid_codigos = ["001", "002", ['003','010']]
        self.conn = DatabaseConnection.get_db_engine(self)
    
    def produtos_query(self, conn, unid_codigo):
        if isinstance(unid_codigo, list):
            unid_values = ",".join([f"'{code}'" for code in unid_codigo])
        else:
            unid_values = f"'{unid_codigo}'"
        query = (f"""
            SELECT 
                prun.prun_unid_codigo AS unidade,
                prun.prun_ativo as tipo,
                prun.prun_prod_codigo AS prod_codigo,
                prod.prod_codbarras AS cod_barras,
                prod.prod_marca AS marca,
                TO_CHAR(prod.prod_codigo, '00999') AS cod_prod,
                forn.forn_razaosocial AS razao_fornecedor,
                forn.forn_cnpjcpf AS cnpj_fornecedor,
                prun.prun_emb AS embalagem,
                prod.prod_descricao AS produto,
                prun.prun_estoque1 AS estoque1,
                prod.prod_status AS status_prod,
                prod.prod_forn_codigo AS cod_fornecedor
                FROM produn AS prun 
                LEFT JOIN produtos AS prod ON prun.prun_prod_codigo = prod.prod_codigo
                LEFT JOIN fornecedores AS forn ON prod.prod_forn_codigo = forn.forn_codigo
                WHERE prun.prun_bloqueado = 'N' 
                AND prun.prun_unid_codigo IN ({unid_values})
                AND prun.prun_ativo = 'S'
                AND prod.prod_marca IN ('BRF','BRF IN NATURA')
                AND prod.prod_status = 'N'
        """)
        return pd.read_sql_query(query, conn)

    def process_rows(self, df, unid_codigo):
        processed_rows = []
        for index, row in df.iterrows():
                        
            registro = 'V'
            # cnpj_forn = row["cnpj_fornecedor"].zfill(14) if row["cnpj_fornecedor"] else '0'*14
            cnpj_forn = ('1838723010513')[:18].ljust(18)
            # razao_forn = (row["razao_fornecedor"] or "")[:30].ljust(30)
            razao_forn = ('BRF S.A.')[:30].ljust(30)
            cod_produto = (row["cod_prod"] or "")[:14].ljust(14)
            
            if (row["embalagem"] != "KG"):
                tipo_emb = "1"
            else:
                tipo_emb = "0"
            
            cod_barras = row["cod_barras"].zfill(14) if row["cod_barras"] else '0'*14
            if cod_barras[0] != '7':
                cod_barras_tipo = "3"
            else:
                cod_barras_tipo = "1"
            
            nome_produto = (row["produto"] or "")[:100].ljust(100)
            razao_forn2 = (row["razao_fornecedor"] or "")[:40].ljust(40)
            estoque_qtde = row["estoque1"]
            status_prod = "A" if estoque_qtde > 0 else "I"

            espaco_branco1 = ' '*30
            espaco_branco2 = ' '*27
            
            processed_row = (f'{registro}{cnpj_forn}{razao_forn}{cod_produto}{tipo_emb}{cod_barras}{cod_barras_tipo}{nome_produto}{razao_forn2}{espaco_branco1}{status_prod}{espaco_branco2}')
            
            processed_rows.append(processed_row)
        logger.info('Query produtos OK!')
        logger.info('Processamento de dados produtos OK!')
        return processed_rows
    
    def save_to_excel_and_txt(self, processed_rows, unid_codigo, data_atual):
        wb = openpyxl.Workbook()
        ws = wb.active
        data_titulo = datetime.now().strftime("%Y%m%d")
        caracter_adc = 'I'
        if unid_codigo == '001':
            cnpj_unid = '81894255000147'
        elif unid_codigo == '002':
            cnpj_unid = '81894255000228'
        else:
            cnpj_unid = '81894255000309'
        ws['A1'] = (f'HCADPROD   {data_titulo}')
        ws['A2'] = (f'{caracter_adc}{cnpj_unid}')
        for index, row_value in enumerate(processed_rows, start=2):
            ws.cell(row=index, column=1).value = row_value
        
        if unid_codigo == ['003','010']:
            unid_codigo = '003'
        nome_arquivo = (f'PRODUTOSDUSNEI{unid_codigo}{data_atual}')
        ws.title = data_atual
        data_pasta = datetime.now().strftime("%Y-%m-%d")
        diretorio = f'{self.path_dados}/{data_pasta}'
        if not os.path.exists(diretorio):
                os.mkdir(diretorio)
        local_arquivo = os.path.join(f'{diretorio}/{nome_arquivo}.xlsx')
        wb.save(local_arquivo)
        
        time.sleep(2)
        
        local_arquivo_txt = os.path.join(f'{diretorio}/{nome_arquivo}.txt')
        with open(local_arquivo_txt, 'w') as txt_file:
            txt_file.write((f'HCADPROD   {data_titulo}') + '\n')
            txt_file.write((f'{caracter_adc}{cnpj_unid}') + '\n')
            for row in processed_rows:
                txt_file.write(row + '\n')
        
        logger.info('Arquivo produtos OK!')
    
    def produtos(self):
        for unid_codigo in self.unid_codigos:
            df = pd.concat([self.produtos_query( self.conn, unid_codigo)])
            
            processed_rows = self.process_rows(df, unid_codigo)
            data_atual = datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]
            self.save_to_excel_and_txt(processed_rows, unid_codigo, data_atual)
        
        logger.info('Funções produtos OK!')
            
if __name__ == "__main__":
    db_clientes = Produtos()    
    db_clientes.produtos()
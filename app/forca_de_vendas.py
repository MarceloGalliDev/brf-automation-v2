# flake8: noqa W293
import os
import time
import openpyxl
import pandas as pd
from loguru import logger
from datetime import datetime
from dotenv import load_dotenv
from conn import DatabaseConnection
from unidecode import unidecode

class Forca_De_Venda:
    def __init__(self):
        load_dotenv()
        self.path_dados = os.getenv('DATA_DIRECTORY')
        self.unid_codigos = ["001", "002", ['003','010']]
        self.conn = DatabaseConnection.get_db_engine(self)
    
    def forca_de_vendas_query(self, conn, unid_codigo):
        if isinstance(unid_codigo, list):
            unid_values = ",".join([f"'{code}'" for code in unid_codigo])
        else:
            unid_values = f"'{unid_codigo}'"
        query = (f"""
            SELECT 
                clie.clie_unid_codigo AS unidade,
                clie.clie_codigo AS clie_codigo,
                clie.clie_nome AS clie_nome,
                clie.clie_cnpjcpf AS cnpj_cpf,
                vend.vend_supe_codigo AS supe_codigo,
                LPAD(clie.clie_vend_codigo, 4, '0') AS cod_vendedor,
                REPLACE(UPPER(vend.vend_nome),' ','') AS nome_vendedor,
                TO_CHAR(vend.vend_codfrente, '0000') AS cod_gerente,
                REPLACE(UPPER(vend.vend_extra8),' ','') AS nome_gerente,
                TO_CHAR(supe.supe_codigo::numeric, '0000') AS cod_supervisor,
                REPLACE(UPPER(supe.supe_nome),' ','') AS nome_supervisor
            FROM clientes AS clie 
            LEFT JOIN vendedores AS vend ON clie.clie_vend_codigo = vend.vend_codigo
            LEFT JOIN supervisores AS supe ON vend.vend_supe_codigo = supe.supe_codigo
            LEFT JOIN municipios AS muni ON clie.clie_muni_codigo_res = muni.muni_codigo
            WHERE clie.clie_tipos NOT IN ('','VE','FU','UN','NL')
            AND clie.clie_endres NOT IN ('') 
            AND muni.muni_nome NOT IN ('','IDENTIFICAR', 'Identificar') 
            AND clie.clie_rota_codigo NOT IN ('') 
            AND clie.clie_unid_codigo IN ({unid_values})
            AND clie.clie_cnpjcpf > '0'
            AND clie.clie_cnpjcpf <> ''
            AND clie.clie_cepres NOT IN ('00000-000','','0','00000','00000000')
            AND clie.clie_cepres NOT IN ('')
        """)
        return pd.read_sql_query(query, conn)

    def process_rows(self, df, unid_codigo):
        processed_rows = []
        for index, row in df.iterrows():
            caracter_adc = "D"
            
            if unid_codigo == '001':
                cnpj_unid = '81894255000147'
            elif unid_codigo == '002':
                cnpj_unid = '81894255000228'
            else:
                cnpj_unid = '81894255000309'
            
            cnpj_cliente = row["cnpj_cpf"].zfill(14)
            nome_gerente = row["nome_gerente"]
            if nome_gerente is not None:
                nome_gerente = unidecode(row["nome_gerente"])[:50].ljust(50).upper()
            
            nome_supervisor = row["nome_supervisor"]
            if nome_gerente is not None:
                nome_supervisor = unidecode(row["nome_supervisor"])[:50].ljust(50).upper()
            
            nome_vendedor = row["nome_vendedor"]
            if nome_gerente is not None:
                nome_vendedor = unidecode(row["nome_vendedor"])[:50].ljust(50).upper()
            
            #verificando se não é null o resultado
            cod_gerente = row["cod_gerente"]
            if cod_gerente is not None:
                cod_gerente = unidecode(row["cod_gerente"])[:13].ljust(13)

            cod_supervisor = row["cod_supervisor"]
            if cod_supervisor is not None:
                cod_supervisor = unidecode(row["cod_supervisor"])[:13].ljust(13)
            
            cod_vendedor = row["cod_vendedor"]
            if cod_vendedor is not None:
                cod_vendedor = unidecode(row["cod_vendedor"])[:20].ljust(20)
            
            espaco_branco1 = ' '*4
            
            processed_row = (f'{caracter_adc}{cnpj_unid}{cnpj_cliente}{espaco_branco1}{cod_gerente}{nome_gerente}{cod_supervisor}{nome_supervisor}{cod_vendedor}{nome_vendedor}')
            processed_rows.append(processed_row)
        logger.info('Query estoques OK!')
        logger.info('Processamento de dados estoques OK!')
        return processed_rows
    
    def save_to_excel_and_txt(self, processed_rows, unid_codigo, data_atual):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = ('HFV10   01838723010513')
        for index, row_value in enumerate(processed_rows, start=2):
            ws.cell(row=index, column=1).value = row_value
        
        if unid_codigo == ['003','010']:
            unid_codigo = '003'
        nome_arquivo = (f'FORCAVENDASDUSNEI{unid_codigo}{data_atual}')
        ws.title = data_atual
        data_pasta = datetime.now().strftime("%Y-%m-%d")
        diretorio = f'{self.path_dados}/{data_pasta}'
        if not os.path.exists(diretorio):
                os.mkdir(diretorio)
        local_arquivo = os.path.join(f'{diretorio}/{nome_arquivo}.xlsx')
        wb.save(local_arquivo)
        
        time.sleep(2)
        
        local_arquivo_txt = os.path.join(f'{diretorio}/{nome_arquivo}.txt')
        with open(local_arquivo_txt, 'w') as txt_file:
            txt_file.write('HFV10   01838723010513' + '\n')
            for row in processed_rows:
                txt_file.write(row + '\n')
        
        logger.info('Arquivo forca_de_vendas OK!')
    
    def forca_de_vendas(self):
        for unid_codigo in self.unid_codigos:
            df = pd.concat([self.forca_de_vendas_query( self.conn, unid_codigo)])
            
            processed_rows = self.process_rows(df, unid_codigo)
            data_atual = datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]
            self.save_to_excel_and_txt(processed_rows, unid_codigo, data_atual)
        
        logger.info('Funções forca_de_vendas OK!')
            
if __name__ == "__main__":
    db_clientes = Forca_De_Venda()    
    db_clientes.forca_de_vendas()
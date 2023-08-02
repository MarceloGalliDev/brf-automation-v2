# flake8: noqa W293
import os
import time
import openpyxl
import pandas as pd
from loguru import logger
from datetime import datetime
from dotenv import load_dotenv

import sys
sys.path.append('/app/config/')
from ..config.conn import DatabaseConnection


class Vendas:
    def __init__(self):
        load_dotenv()
        self.path_dados = os.getenv('DATA_DIRECTORY')
        self.unid_codigos = ["001", "002", ['003', '010']]
        self.conn = DatabaseConnection.get_db_engine(self)
   
    def vendas_query(self, table_name, conn, unid_codigo):
        if isinstance(unid_codigo, list):
            unid_values = ",".join([f"'{code}'" for code in unid_codigo])
        else:
            unid_values = f"'{unid_codigo}'"
        query = (f"""
            SELECT 
                mprd.mprd_transacao AS transacao,
                clie.clie_cnpjcpf AS cnpj_cpf,
                mprd.mprd_datamvto AS data_mvto,
                mprd.mprd_numerodcto AS nfe,
                prod.prod_codbarras AS cod_barras,
                prod.prod_pesoliq AS peso,
                prod.prod_codigo AS cod_prod,
                TO_CHAR(mprd.mprd_qtde, '00000000000999D9999') AS qtde,
                TO_CHAR((mprd.mprd_valor / mprd.mprd_qtde), '00999D99') AS valor_unitario_2,
                TO_CHAR(mprd.mprd_vlrunitario, '00999D99') AS valor_unitario,
                mprc.mprc_vend_codigo AS cod_vend,
                mprd.mprd_dcto_codigo AS dcto_cod,
                mprd.mprd_embcom AS embalagem_vend,
                SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
            FROM {table_name} AS mprd
            LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
            LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
            LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
            WHERE mprd_status = 'N'
            AND mprd_unid_codigo IN ({unid_values})
            AND prod.prod_marca IN ('BRF', 'BRF IN NATURA')
            AND mprd.mprd_dcto_codigo IN ('6666','6668','7339','7335','7338','7337','7260','7263','7262','7268','7264','7269','7267','7319','7318', '6680','6890')
            AND mprc.mprc_vend_codigo NOT IN ('0','00','000','0000','')
            AND mprd.mprd_vlrunitario NOT IN ('0.0100000000')
            AND clie.clie_cepres NOT IN ('00000-000','','0','00000','00000000')
            AND clie.clie_cepres > '0'
            AND clie.clie_cepres NOT IN ('')
        """)
            # AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '7 DAYS'
        return pd.read_sql_query(query, conn)
        
    def process_rows(self, df, unid_codigo):
        processed_rows = []
        for index, row in df.iterrows():
            caracter_adc = "D"
            
            if unid_codigo == '001':
                cnpj_unid = '81894255000147'
            elif unid_codigo == '002':
                cnpj_unid = '81894255000228'
            else:
                cnpj_unid = '81894255000309'
                
            cnpj_cliente = row['cnpj_cpf'].zfill(14)
            data_mvto = row['data_mvto'].strftime('%Y%m%d')
            documento_nfe = row['nfe'].zfill(6) 
            cod_barras = row['cod_barras'].zfill(13)
            dcto_cod = row['dcto_cod']
            
            if dcto_cod in ['6666', '6668', '7339', '7335', '7338', '7337']:
                qtde = '0' + row['qtde'].strip()
            else:
                qtde = '-' + row['qtde'].strip()
            
            valor_unitario = (
                row['valor_unitario_2'].strip() 
                if row['valor_unitario'].strip() == '00000.00'
                else row['valor_unitario'].strip()
            )
            
            cod_vend = row['cod_vend'].zfill(4)            
            tipo_doc = 'B' if dcto_cod in ['6890', '7267'] else 'N'
            cep = row['cep']
            tipo_unid = '0002' if row['embalagem_vend'] == 'KG' else '0001'
                            
            espaco_branco1 = ' '*4
            espaco_branco2 = ' '*13
            espaco_branco3 = ' '
            espaco_branco4 = ' '*16
            espaco_branco5 = ' '*10
            espaco_branco6 = ' '*13
            espaco_branco7 = ' '*6
            espaco_branco8 = ' '*2
            espaco_branco9 = ' '
            espaco_branco10 = '00000.00'
            
            processed_row = (f'{caracter_adc}{cnpj_unid}{cnpj_cliente}{espaco_branco1}{data_mvto}{documento_nfe}{espaco_branco2}{cod_barras}{espaco_branco3}{qtde}{valor_unitario}{cod_vend}{espaco_branco4}{espaco_branco5}{tipo_doc}{cep}{espaco_branco6}{espaco_branco7}{espaco_branco8}{espaco_branco9}{espaco_branco10}{tipo_unid}') # noqa: E501, E261
            processed_rows.append(processed_row)
            
        logger.info('Query vendas OK!')
        logger.info('Processamento de dados vendas OK!')
        return processed_rows
    
    def save_to_excel_and_txt(self, processed_rows, unid_codigo, data_atual):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'HVENDA1101838723010513'
        for index, row_value in enumerate(processed_rows, start=2):
            ws.cell(row=index, column=1).value = row_value
            
        if unid_codigo == ['003','010']:
            unid_codigo = '003'
        nome_arquivo = (f'VENDASDUSNEI{unid_codigo}{data_atual}')
        ws.title = data_atual
        data_pasta = datetime.now().strftime("%Y-%m-%d")
        diretorio = f'{self.path_dados}/{data_pasta}'
        if not os.path.exists(diretorio):
            os.mkdir(diretorio)
        local_arquivo = os.path.join(f'{diretorio}/{nome_arquivo}.xlsx')
        wb.save(local_arquivo)
        
        time.sleep(5)
        
        local_arquivo_txt = os.path.join(f'{diretorio}/{nome_arquivo}.txt')
        with open(local_arquivo_txt, 'w') as txt_file:
            txt_file.write('HVENDA1101838723010513' + '\n')
            for row in processed_rows:
                txt_file.write(row + '\n')
        
        logger.info('Arquivo vendas OK!')
        
        
    def vendas(self):
        for unid_codigo in self.unid_codigos:
            tables = ['movprodd0122', 'movprodd0222', 'movprodd0322', 'movprodd0422', 'movprodd0522', 'movprodd0622', 'movprodd0722', 'movprodd0822', 'movprodd0922', 'movprodd1022', 'movprodd1122', 'movprodd1222', 'movprodd0123', 'movprodd0223', 'movprodd0323', 'movprodd0423', 'movprodd0523', 'movprodd0623', 'movprodd0723', 'movprodd0823', 'movprodd0923', 'movprodd1023',  'movprodd1123', 'movprodd1223']
            
            df = pd.concat([self.vendas_query(table, self.conn, unid_codigo)for table in tables])
            
            processed_rows = self.process_rows(df, unid_codigo)
            data_atual = datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]
            self.save_to_excel_and_txt(processed_rows, unid_codigo, data_atual)
        
        logger.info('Funções vendas OK!')
     
            
if __name__ == "__main__":
    db_vendas = Vendas()
    db_vendas.vendas()

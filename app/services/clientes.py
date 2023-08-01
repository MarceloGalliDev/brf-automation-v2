import os
import time
import openpyxl
import pandas as pd
from loguru import logger
from datetime import datetime
from dotenv import load_dotenv
from config import conn
from unidecode import unidecode

class Clientes:
    def __init__(self):
        load_dotenv()
        self.path_dados = os.getenv('DATA_DIRECTORY')
        self.unid_codigos = ['001', '002', ['003','010']]
        self.conn = conn.DatabaseConnection.get_db_engine(self)
    
    def clientes_query(self, conn, unid_codigo):
        if isinstance(unid_codigo, list):
            unid_values = ",".join([f"'{code}'" for code in unid_codigo])
        else:
            unid_values = f"'{unid_codigo}'"
        query = (f"""
            SELECT 
                clie.clie_unid_codigo AS unidade,
                clie.clie_codigo AS clie_codigo,
                REPLACE(UPPER(clie.clie_nome),',','') AS clie_nome,
                clie.clie_cnpjcpf AS cnpj_cpf,
                REPLACE(UPPER(LEFT(clie.clie_razaosocial, 40)),' ','') AS razao_social,
                REPLACE(UPPER(clie.clie_endres),' ','') AS endereco,
                REPLACE(UPPER(clie.clie_endresnumero),' ','') AS numero_res,
                REPLACE(REPLACE(UPPER(clie.clie_bairrores),' ',''),',','') AS bairro,
                SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep,
                clie.clie_muni_codigo_res AS code_muni,
                muni.muni_codigo AS cod_municipio,
                REPLACE(UPPER(muni.muni_nome),' ','') AS municipio,
                clie.clie_ufexprg AS estado,
                clie.clie_rota_codigo AS rota,
                clie.clie_ramoatividade AS ramo
            FROM clientes AS clie 
            LEFT JOIN municipios AS muni ON clie.clie_muni_codigo_res = muni.muni_codigo
            WHERE clie.clie_tipos NOT IN ('','VE','FU','UN','NL')
            AND clie.clie_endres NOT IN ('') 
            AND muni.muni_nome NOT IN ('','IDENTIFICAR', 'Identificar') 
            AND clie.clie_rota_codigo NOT IN ('') 
            AND clie.clie_unid_codigo IN ({unid_values})
            AND clie.clie_unid_codigo NOT IN ('')
            AND clie.clie_cnpjcpf > '0'
            AND clie.clie_cepres NOT IN ('00000-000','','0','00000','00000000')
            AND clie.clie_cepres > '0'
            AND clie.clie_cepres NOT IN ('')
            AND clie.clie_razaosocial NOT IN ('TESTE', '')
        """)
        return pd.read_sql_query(query, conn)

    def process_rows(self, df, unid_codigo):
        processed_rows = []
        for index, row in df.iterrows():
            caracter_adc = "D"
            
            if unid_codigo == '001':
                cnpj_unid = '81894255000147'
            elif unid_codigo == '002':
                cnpj_unid = '81894255000228'
            else:
                cnpj_unid = '81894255000309'
            
            cnpj_cliente = row['cnpj_cpf'].zfill(14)
            razao_social = unidecode(row['razao_social'])[:40].ljust(40).upper()
            endereco = unidecode(row['endereco']).upper()
            numero_res = unidecode(row['numero_res'])
            endereco_num = unidecode(f'{endereco}{numero_res}')[:40].ljust(40)
            bairro = unidecode(row['bairro'])[:30].ljust(30).upper()
            cep = row['cep']
            cidade = unidecode(row['municipio'])[:30].ljust(30).upper()
            estado = unidecode(row['estado'])[:30].ljust(30).upper()
            rota = row['rota'][:10].ljust(10) 
            segmentos = unidecode('OUTROS')[:10].ljust(10)
            
            espaco_branco1 = ' '*4
            espaco_branco2 = ' '*20
            espaco_branco3 = ' '*40
            espaco_branco4 = ' '*10
            espaco_branco5 = ' '*6
            
            processed_row = (f'{caracter_adc}{cnpj_unid}{cnpj_cliente}{espaco_branco1}{razao_social}{endereco_num}{bairro}{cep}{cidade}{estado}{espaco_branco2}{espaco_branco3}{cnpj_cliente}{espaco_branco1}{rota}{espaco_branco4}{segmentos}{espaco_branco5}')
            
            processed_rows.append(processed_row)
        logger.info('Query clientes OK!')
        logger.info('Processamento de dados clientes OK!')
        return processed_rows

    def save_to_excel_and_txt(self, processed_rows, unid_codigo, data_atual):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = (f'HPDV10  01838723010513')
        for index, row_value in enumerate(processed_rows, start=2):
            ws.cell(row=index, column=1).value = row_value
        
        if unid_codigo == ['003','010']:
            unid_codigo = '003'
        nome_arquivo = (f'CLIENTESDUSNEI{unid_codigo}{data_atual}')
        ws.title = data_atual
        data_pasta = datetime.now().strftime("%Y-%m-%d")
        diretorio = f'{self.path_dados}/{data_pasta}'
        if not os.path.exists(diretorio):
                os.mkdir(diretorio)
        local_arquivo = os.path.join(f'{diretorio}/{nome_arquivo}.xlsx')
        wb.save(local_arquivo)
        
        time.sleep(5)
        
        local_arquivo_txt = os.path.join(f'{diretorio}/{nome_arquivo}.txt')
        with open(local_arquivo_txt, 'w') as txt_file:
            txt_file.write('HPDV10  01838723010513' + '\n')
            for row in processed_rows:
                txt_file.write(row + '\n')
        
        logger.info('Arquivo clientes OK!')
    
    def clientes(self):
        for unid_codigo in self.unid_codigos:
            df = pd.concat([self.clientes_query( self.conn, unid_codigo)])
            
            processed_rows = self.process_rows(df, unid_codigo)
            data_atual = datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]
            self.save_to_excel_and_txt(processed_rows, unid_codigo, data_atual)
        
        logger.info('Funções clientes OK!')
            
if __name__ == "__main__":
    db_clientes = Clientes()    
    db_clientes.clientes()